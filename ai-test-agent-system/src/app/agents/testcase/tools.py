"""测试用例Agent的工具定义。

此模块包含所有可用的工具定义，包括：
- 基础工具：导出测试用例到Excel
- RAG工具：通过MCP客户端获取的检索增强工具
"""

import asyncio
from functools import lru_cache
from typing import Union

from langchain.tools import tool
from langchain_core.tools import BaseTool
from langchain_mcp_adapters.client import MultiServerMCPClient
from app.processors.pdf import extract_pdf_text_from_file

# MCP服务器配置
MCP_SERVER_CONFIGS = {
    "rag-server": {
        # SSE 服务端点 URL
        "url": "http://localhost:8008/sse",
        # 传输协议：sse (Server-Sent Events)
        "transport": "sse",
    }
}

"""
测试用例 Excel 导出工具

本模块提供将测试用例导出为 Excel 文件的能力，支持企业级测试管理工具
（如禅道、Tapd、TestRail）的导入格式。
"""

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# Excel 样式配置
_HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_ALIGNMENT_WRAP = Alignment(vertical="top", wrap_text=True)
_ALIGNMENT_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 默认列宽配置
_DEFAULT_COLUMN_WIDTHS = {
    "A": 18,  # 用例编号
    "B": 35,  # 用例标题
    "C": 14,  # 所属模块
    "D": 12,  # 用例类型
    "E": 10,  # 优先级
    "F": 30,  # 前置条件
    "G": 40,  # 测试步骤
    "H": 30,  # 测试数据
    "I": 40,  # 预期结果
    "J": 20,  # 备注
}


def _flatten_steps(steps: list[dict[str, Any]] | None) -> str:
    """将步骤列表转换为带序号的文本。"""
    if not steps:
        return ""
    lines = []
    for step in steps:
        seq = step.get("seq", step.get("step", len(lines) + 1))
        action = step.get("action", step.get("操作描述", ""))
        target = step.get("target", step.get("操作对象", ""))
        data = step.get("data", "")
        line = f"{seq}. {action}"
        if target:
            line += f" [{target}]"
        if data:
            line += f"（数据：{data}）"
        lines.append(line)
    return "\n".join(lines)


def _flatten_test_data(test_data: dict[str, Any] | str | None) -> str:
    """将测试数据转换为文本。"""
    if not test_data:
        return ""
    if isinstance(test_data, str):
        return test_data
    lines = [f"{k}: {v}" for k, v in test_data.items()]
    return "\n".join(lines)


def _flatten_expected_results(expected_results: list[str] | str | None) -> str:
    """将预期结果列表转换为文本。"""
    if not expected_results:
        return ""
    if isinstance(expected_results, str):
        return expected_results
    lines = []
    for idx, result in enumerate(expected_results, start=1):
        lines.append(f"{idx}. {result}")
    return "\n".join(lines)


def _flatten_preconditions(preconditions: list[str] | str | None) -> str:
    """将前置条件列表转换为文本。"""
    if not preconditions:
        return ""
    if isinstance(preconditions, str):
        return preconditions
    lines = []
    for idx, cond in enumerate(preconditions, start=1):
        lines.append(f"{idx}. {cond}")
    return "\n".join(lines)


def _extract_field(case: dict[str, Any], *keys: str, default: Any = "") -> Any:
    """从字典中按多个候选键提取值。"""
    for key in keys:
        if key in case:
            return case[key]
    return default

@tool
def export_test_cases_to_excel(
    test_cases: list[dict[str, Any]],
    output_path: str | Path,
    sheet_name: str = "测试用例",
) -> str:
    """
    将测试用例列表导出为 Excel 文件。

    支持的测试用例字段（兼容 JSON / CSV / Markdown 中定义的格式）：
      - id / 用例编号
      - title / 用例标题
      - module / 所属模块
      - type / 用例类型
      - priority / 优先级
      - preconditions / 前置条件
      - steps / 测试步骤
      - test_data / 测试数据
      - expected_results / 预期结果
      - remarks / 备注

    Args:
        test_cases: 测试用例字典列表，每个字典描述一条用例。
        output_path: 导出的 Excel 文件路径（支持 str 或 Path）。
        sheet_name: 工作表名称，默认为 "测试用例"。

    Returns:
        导出文件的绝对路径字符串。
    """
    if not test_cases:
        raise ValueError("测试用例列表为空，无法导出 Excel。")

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("无法创建工作表。")
    ws.title = sheet_name

    # 表头
    headers = [
        "用例编号",
        "用例标题",
        "所属模块",
        "用例类型",
        "优先级",
        "前置条件",
        "测试步骤",
        "测试数据",
        "预期结果",
        "备注",
    ]
    ws.append(headers)

    # 设置表头样式
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _ALIGNMENT_CENTER
        cell.border = _BORDER

    # 写入数据行
    for case in test_cases:
        row = [
            _extract_field(case, "id", "用例编号"),
            _extract_field(case, "title", "用例标题"),
            _extract_field(case, "module", "所属模块"),
            _extract_field(case, "type", "用例类型"),
            _extract_field(case, "priority", "优先级"),
            _flatten_preconditions(_extract_field(case, "preconditions", "前置条件", default=None)),
            _flatten_steps(_extract_field(case, "steps", "测试步骤", default=None)),
            _flatten_test_data(_extract_field(case, "test_data", "测试数据", default=None)),
            _flatten_expected_results(_extract_field(case, "expected_results", "预期结果", default=None)),
            _extract_field(case, "remarks", "备注"),
        ]
        ws.append(row)
        row_idx = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = _ALIGNMENT_WRAP
            cell.border = _BORDER

    # 设置列宽
    for col_letter, width in _DEFAULT_COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # 设置行高（给内容较多的行留出空间）
    ws.row_dimensions[1].height = 24
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 60

    wb.save(str(output_path))
    return str(output_path.resolve())


@lru_cache(maxsize=1)
def _cached_rag_tools() -> tuple[BaseTool, ...]:
    """缓存RAG工具列表，避免重复创建MCP客户端。
    
    Returns:
        RAG工具的元组（不可变，可缓存）
    """
    client = MultiServerMCPClient({
    "rag-server": {
        # SSE 服务端点 URL
        "url": "http://localhost:8008/sse",
        # 传输协议：sse (Server-Sent Events)
        "transport": "sse",
    }
})
    tools = asyncio.run(client.get_tools())
    return tuple(tools)


def rag_mcp_tools() -> list[BaseTool]:
    """获取RAG MCP工具列表。
    
    通过MCP客户端从远程服务器获取RAG检索工具。
    结果会被缓存以避免重复创建连接。
    
    Returns:
        RAG工具列表
    """
    return list(_cached_rag_tools())


def get_rag_tool_names() -> set[str]:
    """获取RAG工具的名称集合，用于识别和过滤。"""
    return {tool.name for tool in rag_mcp_tools()}


def get_tool_name(tool: Union[BaseTool, dict]) -> str:
    """获取工具名称，支持 BaseTool 对象和字典格式。
    
    Args:
        tool: 工具对象（BaseTool 或 dict）
        
    Returns:
        工具名称字符串
    """
    if isinstance(tool, dict):
        return tool.get("name", "")
    return getattr(tool, "name", "")


# RAG 系统提示词扩展（已精简，详细规范请参见 rag-query Skill）
RAG_SYSTEM_PROMPT_APPENDIX = """

---

## 附录：可用 RAG 工具列表

{rag_tools_description}

> 详细的 RAG 检索策略、mode 选择规范、结果引用规范等，请严格遵循 `rag-query` Skill 执行。
"""


def format_rag_tools_description() -> str:
    """格式化RAG工具描述，用于系统提示词。
    
    Returns:
        RAG工具描述文本
    """
    tools = rag_mcp_tools()
    if not tools:
        return "（暂无RAG工具配置）"
    
    descriptions = []
    for tool in tools:
        desc = getattr(tool, 'description', '无描述')
        descriptions.append(f"- **{tool.name}**: {desc}")
    return "\n".join(descriptions)

def get_next_ai_drawio_tools():
    client = MultiServerMCPClient({
    "drawio": {
      "command": "npx",
      "args": ["@next-ai-drawio/mcp-server@latest"],
      "transport": "stdio",
        }
    })
    tools = asyncio.run(client.get_tools())
    return tools

def get_all_tools() -> list:
    """获取所有可用工具的完整列表。
    
    包括基础工具和RAG工具，用于在 create_agent 中注册。
    
    Returns:
        所有工具的列表
    """
    return [export_test_cases_to_excel, extract_pdf_text_from_file] + rag_mcp_tools()


def get_base_tools() -> list:
    """获取基础工具列表（不包含RAG工具）。
    
    Returns:
        基础工具列表
    """
    return [export_test_cases_to_excel, extract_pdf_text_from_file] # + get_next_ai_drawio_tools()
